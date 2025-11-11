from defect_list_xslx_file_reader import get_defect_data_from_sheet
from torex_xlsb_file_reader import get_torex_data_from_sheet
from valve_type_search_in_torex import valve_type_search_in_torex
from valve_type_analyzer import get_type_from_element_name, filter_valves_by_name
from defect_type_analyzer import get_defect_type, get_defect_class
from calculate_defect_intensivity import calculate_defect_intensivity
from model.valve import Valve, Defect
from model.torex_record import TorexRecord
from json import load
from typing import Dict, List, Tuple
import openpyxl
import os
from scripts.get_analyzis_for_filtered_1 import get_analyzis_for_filtered_1
from scripts.get_analyzis_for_filtered_2 import get_analyzis_for_filtered_2
from scripts.get_analyzis_for_filtered_3_with_VAB import get_analyzis_for_filtered_3
from scripts.statistic_analyzis_for_manufacturers_specified_and_type import *
from scripts.select_valves_by_type import select_valves_by_type

# ===== КОНФИГУРАЦИЯ =====
excel_file_path = (
    "files\\данные\\Карта данных 2015-2025.xlsx"  # Укажите путь к вашему файлу
)
torex_file_path = (
    "files\\данные\\Регистр ТОРЭКС_прореженный.xlsx"  # Укажите путь к вашему файлу
)
types_blacklist_path = "files\\settings\\names_blacklist.txt"
types_whitelist_path = "files\\settings\\names_whitelist.json"
defect_types_path = "files\\settings\\defect_list.json"
defect_classes_path = "files\\settings\\defect_classes.json"
manufacturers_path = "files\\settings\\manufacturers.json"
column_names_path = "column_names.json"
column_names_torex_path = "column_names_torex.json"
result_folder_path = "files\\result"
result_manufacturer_folder_path = "files\\result_manufacturer"
output_manufacturer_folder_path = "files\\output_manufacturer"
result_manufacturer_defined_folder_path = "files\\result_manufacturer_defined"
result_manufacturer_valve_type_folder_path = "files\\result_manufacturer_valve_type_new"
output_manufacturer_defined_folder_path = "files\\output_manufacturer_defined"
NUMBER_OF_INTERVALS = 10  # Количество интервалов разбиения
NUMBER_OF_ITEMS = 2181  # Общее количество наблюдаемых изделий
CONFIDENCE_LEVEL = 0.9  # Уровень доверия (P=0.9)


vab = {
    "3-4": (2.73 + 8.95) * 24 * (10 ** (-7)),
    "5": (3.0 + 1.78) * 24 * (10 ** (-6)),
    "6-7": (3.0 + 1.78) * 24 * (10 ** (-6)),
}
# =========================


def check_valve_type(valve: Valve):
    result = False
    if valve.valve_type is not None and valve.main_valve_type is not None:
        if valve.valve_type.upper().find(valve.main_valve_type.upper()) == -1 or (
            valve.get_descritpion() != ""
            and valve.get_descritpion().find(valve.main_valve_type.upper()) == -1
        ):
            result = True
    return result


def get_manufacturer_from_valve(valve: Valve, whitelist: Dict[str, str]):
    result = False
    for manufacturer in whitelist:
        for manufacturer_name_variant in whitelist[manufacturer]:
            if valve.manufacturer.upper().find(manufacturer_name_variant.upper()) != -1:
                result = True
                valve.manufacturer_defined = manufacturer
                return result
    return result


def create_file(path):
    directory_path = os.path.dirname(path)
    if not os.path.isdir(directory_path):
        os.makedirs(directory_path, exist_ok=True)
    return open(path, "w", encoding="utf-8")


def append_to_file(path, new_str):
    directory_path = os.path.dirname(path)
    if not os.path.isdir(directory_path):
        os.makedirs(directory_path, exist_ok=True)
    with open(path, "a", encoding="utf-8") as file_to_append:
        print(new_str, file=file_to_append)


types_blacklist: List[str] = []

with open(types_blacklist_path, encoding="utf-8") as types_blacklist_file:
    for line in types_blacklist_file.readlines():
        types_blacklist.append(line.strip())
with open(types_whitelist_path, encoding="utf-8") as whitelist_file:
    types_whitelist = load(whitelist_file)
with open(defect_types_path, encoding="utf-8") as defect_types_file:
    defect_types = load(defect_types_file)
with open(column_names_path, encoding="utf-8") as column_names_file:
    column_names = load(column_names_file)
with open(column_names_torex_path, encoding="utf-8") as column_names_file:
    column_names_torex = load(column_names_file)
with open(manufacturers_path, encoding="utf-8") as manufacturers_file:
    manufacturers_list = load(manufacturers_file)
with open(defect_classes_path, encoding="utf-8") as defect_classes_file:
    defect_classes = load(defect_classes_file)

workbook = openpyxl.load_workbook(excel_file_path)
sheet_names = workbook.sheetnames
defect_list_file_data: List[List[Valve]] = []
for sheet_name in sheet_names:
    sheet = workbook[sheet_name]
    defect_list_file_data.append(get_defect_data_from_sheet(sheet, column_names))

valve_list = {
    f"{valve.get_block_number()}_{valve.get_valve_name_upper()}": valve
    for valve in defect_list_file_data[0]
}

workbook = openpyxl.load_workbook(torex_file_path)
sheet_names = workbook.sheetnames
torex_file_data: List[List[TorexRecord]] = []
for sheet_name in sheet_names:
    sheet = workbook[sheet_name]
    torex_file_data.append(get_torex_data_from_sheet(sheet, column_names_torex))

torex_list = {
    f"{torex_record.get_block_number()}_{torex_record.get_valve_name_operational_upper()}": torex_record
    for torex_record in torex_file_data[0]
}
torex_errors_file = open("files//torex_errors.csv", "w", encoding="utf-8")
for record in torex_list:
    if torex_list[record].valve_name_plant != torex_list[record].valve_name_operational:
        print(torex_list[record], file=torex_errors_file)

valve_type_search_in_torex(valve_list, torex_list)


for valve_name in valve_list:
    get_type_from_element_name(valve_list[valve_name], types_whitelist)
    get_manufacturer_from_valve(valve_list[valve_name], manufacturers_list)
    get_defect_type(valve_list[valve_name], defect_types)
    get_defect_class(valve_list[valve_name], defect_classes)

# select_valves_by_type(valve_list, defect_types, types_whitelist)

# get_analyzis_for_filtered_1(
#     valve_list, NUMBER_OF_INTERVALS, NUMBER_OF_ITEMS, CONFIDENCE_LEVEL
# )
# get_analyzis_for_filtered_2(
#     valve_list, NUMBER_OF_INTERVALS, NUMBER_OF_ITEMS, CONFIDENCE_LEVEL
# )

get_analyzis_for_filtered_3(
    valve_list, NUMBER_OF_INTERVALS, NUMBER_OF_ITEMS, CONFIDENCE_LEVEL, vab
)

valve_by_manufacturers_list = {}
for valve_name in valve_list:
    if not valve_list[valve_name].manufacturer_defined in valve_by_manufacturers_list:
        valve_by_manufacturers_list[valve_list[valve_name].manufacturer_defined] = {}
        
    valve_by_manufacturers_list[valve_list[valve_name].manufacturer_defined][valve_name] = (valve_list[valve_name])
statistic_analyzis_for_manufacturers_specified_and_type(
    valve_by_manufacturers_list,
    result_manufacturer_valve_type_folder_path,
    NUMBER_OF_INTERVALS,
    NUMBER_OF_ITEMS,
    CONFIDENCE_LEVEL,
    vab,
)
